"""
Excel export using openpyxl → BytesIO.
"""
import io
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import openpyxl
from openpyxl.utils import get_column_letter

from ..models.property_record import PropertyRecord
from ..models.land_record import LandRecord
from ..models.business_record import BusinessRecord
from ..models.open_data_record import OpenDataRecord

MODEL_MAP = {
    "properties": PropertyRecord,
    "land-records": LandRecord,
    "businesses": BusinessRecord,
    "open-data": OpenDataRecord,
}


async def export_excel(record_type: str, db: AsyncSession) -> bytes:
    Model = MODEL_MAP.get(record_type)
    if not Model:
        raise ValueError(f"Unknown record type: {record_type}")

    columns = [c.key for c in Model.__mapper__.column_attrs]
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title=record_type[:31])

    # Header row
    ws.append(columns)

    # Data rows in chunks
    offset = 0
    chunk = 500
    while True:
        result = await db.execute(
            select(Model).order_by(Model.id).limit(chunk).offset(offset)
        )
        rows = result.scalars().all()
        if not rows:
            break
        for row in rows:
            ws.append([str(getattr(row, c, "") or "") for c in columns])
        offset += chunk

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
